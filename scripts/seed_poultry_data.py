import sqlite3
import openpyxl
import os
import json
from datetime import datetime, timedelta

# Paths
egg_log_path = "/home/kelvitz/Downloads/kenchic_egg_log.xlsx"
db_path = "/home/kelvitz/AntigravityProjects/PoultryProject/data/poultry.db"

os.makedirs(os.path.dirname(db_path), exist_ok=True)

# ─── Constants ─────────────────────────────────────────────────────────────────
BATCH_ID       = "1779692918051"
TOTAL_BIRDS    = 55
ARRIVAL_DATE   = "2026-01-12"   # arrived Jan 12 at ~2 months old
HATCH_DATE     = "2025-11-17"   # estimated hatch (2 months before arrival)
# "startDate" = when the batch cockpit tracks from (arrival = start of this operation)
START_DATE     = "2026-01-12T00:00:00.000Z"
LAY_START_DATE = "2026-05-25"   # first egg recorded
FEED_KG_DAY    = 5.0            # assumed 5 kg/day since onset of laying

# ─── Vaccine schedule ──────────────────────────────────────────────────────────
# Birds arrived Jan 12 already fully vaccinated at ~8 weeks old.
# Kenchic schedule (estimated based on hatch date Nov 17, 2025):
#   Day 1   → Marek's (at hatchery)
#   Day 5-7 → Newcastle HB1/La Sota  → ~Nov 22, 2025
#   Day 10-14 → Gumboro (IBD)        → ~Nov 27, 2025
#   Day 18-21 → Newcastle Booster    → ~Dec 05, 2025
#   Day 24-28 → Gumboro Booster      → ~Dec 11, 2025
#   Week 8    → Newcastle Komarov    → ~Jan 12, 2026 (arrival day)
#   Week 14-16 → Newcastle La Sota (Pre-lay) → ~Feb 23, 2026
VACCINES = [
    { "drug": "Marek's Disease",          "date": "2025-11-17", "method": "Injection (at hatchery)", "notes": "Done by hatchery before pickup" },
    { "drug": "Newcastle (HB1/La Sota)",  "date": "2025-11-22", "method": "Eye drop / Drinking water", "notes": "First dose" },
    { "drug": "Gumboro (IBD)",            "date": "2025-11-27", "method": "Drinking water", "notes": "First dose" },
    { "drug": "Newcastle (Booster)",      "date": "2025-12-05", "method": "Drinking water", "notes": "Second dose" },
    { "drug": "Gumboro (Booster)",        "date": "2025-12-11", "method": "Drinking water", "notes": "Second dose" },
    { "drug": "Newcastle (Komarov)",      "date": "2026-01-12", "method": "Injection (IM)", "notes": "Long-lasting protection — administered on arrival" },
    { "drug": "Newcastle (La Sota)",      "date": "2026-02-23", "method": "Drinking water", "notes": "Pre-lay booster (estimated week 14-16)" },
]

# ─── Parse Egg Log Excel ───────────────────────────────────────────────────────
wb_egg = openpyxl.load_workbook(egg_log_path, data_only=True)
sheet_egg = wb_egg.active

daily_logs = []
for r in range(5, 18):
    date_val = sheet_egg.cell(row=r, column=1).value
    if not date_val:
        continue
    day_num  = int(sheet_egg.cell(row=r, column=2).value or 0)
    eggs     = int(sheet_egg.cell(row=r, column=3).value or 0)
    notes    = sheet_egg.cell(row=r, column=6).value

    # Parse DD/MM/YYYY
    parsed_date = datetime.strptime(str(date_val).strip(), "%d/%m/%Y")
    iso_date = parsed_date.strftime("%Y-%m-%d")

    daily_logs.append({
        "id":       f"{BATCH_ID}_{iso_date}",
        "date":     iso_date,
        "day":      day_num,
        "eggs":     eggs,
        "birds":    TOTAL_BIRDS,
        "mortality": 0,
        "feed":     FEED_KG_DAY,   # field name is "feed" (kg)
        "water":    0,
        "notes":    notes or ""
    })

# ─── Database Setup ────────────────────────────────────────────────────────────
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

cursor.executescript("""
    CREATE TABLE IF NOT EXISTS batches (
        id TEXT PRIMARY KEY, data TEXT NOT NULL,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS logs (
        id TEXT PRIMARY KEY, batch_id TEXT NOT NULL,
        data TEXT NOT NULL, date TEXT NOT NULL,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS transactions (
        id TEXT PRIMARY KEY, batch_id TEXT NOT NULL,
        data TEXT NOT NULL,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS health_logs (
        id TEXT PRIMARY KEY, batch_id TEXT NOT NULL,
        data TEXT NOT NULL,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
""")

# ─── Insert Batch ──────────────────────────────────────────────────────────────
batch_data = {
    "id":         BATCH_ID,
    "name":       "Batch 001 — ISA Brown Layers",
    "type":       "layers",
    "size":       TOTAL_BIRDS,          # <-- frontend reads b.size
    "startDate":  START_DATE,           # <-- frontend reads b.startDate
    "hatchDate":  HATCH_DATE,
    "arrivalDate": ARRIVAL_DATE,
    "status":     "Active",
    "notes":      f"{TOTAL_BIRDS} birds total (8-12 cockerels). Arrived Jan 12 at ~2 months old, fully vaccinated. Egg laying commenced {LAY_START_DATE}.",
    "stats": {
        "birdsAlive": TOTAL_BIRDS,      # <-- frontend reads stats.birdsAlive
        "totalEggs":  sum(l["eggs"] for l in daily_logs),
        "mortality":  0
    }
}
cursor.execute(
    "INSERT OR REPLACE INTO batches (id, data, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
    (BATCH_ID, json.dumps(batch_data))
)

# ─── Insert Daily Egg Logs ─────────────────────────────────────────────────────
for log in daily_logs:
    cursor.execute(
        "INSERT OR REPLACE INTO logs (id, batch_id, data, date, updated_at) VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)",
        (log["id"], BATCH_ID, json.dumps(log), log["date"])
    )

# ─── Insert Feed Purchase Transaction ─────────────────────────────────────────
# 500 kg feed mixed May 9 — cost unknown, record as 0 for now
tx_feed = {
    "id":       f"tx_{BATCH_ID}_feed_mix_01",
    "type":     "purchase",
    "category": "feed",
    "amount":   0,
    "qty":      500,
    "unit":     "kg",
    "notes":    "500kg Commercial Layer Mash formulation mixed. Ingredients: Broken Maize 119kg, Wheat Bran 200kg, Limestone 40kg, De Heus Concentrate 50kg, Havens Concentrate 25kg, Sunflower 26kg, Soya 29kg, Cotton Cake 11kg.",
    "date":     "2026-05-09",
    "status":   "paid"
}
cursor.execute(
    "INSERT OR REPLACE INTO transactions (id, batch_id, data, updated_at) VALUES (?, ?, ?, CURRENT_TIMESTAMP)",
    (tx_feed["id"], BATCH_ID, json.dumps(tx_feed))
)

# ─── Insert Vaccination Health Logs ───────────────────────────────────────────
for i, vax in enumerate(VACCINES):
    vax_id = f"hl_{BATCH_ID}_vax_{i:02d}"
    health_data = {
        "id":     vax_id,
        "type":   "vaccine",
        "drug":   vax["drug"],
        "date":   vax["date"],
        "method": vax["method"],
        "notes":  vax["notes"],
        "offLabel": False
    }
    cursor.execute(
        "INSERT OR REPLACE INTO health_logs (id, batch_id, data, updated_at) VALUES (?, ?, ?, CURRENT_TIMESTAMP)",
        (vax_id, BATCH_ID, json.dumps(health_data))
    )

conn.commit()
conn.close()

total_eggs = sum(l["eggs"] for l in daily_logs)
print(f"✅ Seeded:")
print(f"   1 batch  → {TOTAL_BIRDS} birds, started {ARRIVAL_DATE}")
print(f"   {len(daily_logs)} daily egg logs → {total_eggs} total eggs")
print(f"   1 feed purchase transaction (500kg)")
print(f"   {len(VACCINES)} vaccination health log entries")
